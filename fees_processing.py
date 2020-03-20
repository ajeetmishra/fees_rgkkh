import os, xlrd, re
import numpy as np
import pandas as pd

#os.chdir('c:/users/abamb2o/downloads/rgkkh/')

class fee_file:
    
    def __init__(self, fname):

        self.fname = fname
        self.proc_status=[]

        self.wb = xlrd.open_workbook(self.fname)
        self.ws = self.wb.sheet_by_name('Sheet 1')

        #Get account number
        try:
            self.ac_no = re.findall(r'\d{13,15}', self.ws.cell(14,4).value)[0]
        except:
            self.proc_status.append("Err - AcNo")
        
        #Get account type - One-time, lodging, tuition fee
        if self.ac_no == '50100103043898':
            self.ac_type = 'Lodging Boarding'
        elif self.ac_no == '50100103044174':
            self.ac_type = 'Tuition'
        elif self.ac_no == '50100103044190':
            self.ac_type = 'One time'
        else:
            self.ac_type = 'Unknown'
            self.proc_status.append("Err - A/c Type")

        #Get date range
        try:
            self.date_range = self.ws.cell(15,0).value
            self.date_from = re.findall('\d{2}/\d{2}/\d{4}', self.date_range)[0]
            self.date_to = re.findall('\d{2}/\d{2}/\d{4}', self.date_range)[1]
        except:
            self.proc_status.append("Err - Date Range")

        #Get number of rows
        self.last_row = self.ws.col_values(0).index('STATEMENT SUMMARY  :-') - 5
        self.count_rows = self.last_row - 21

        self.df = pd.read_excel(self.fname, skiprows=21, nrows=self.count_rows)
        self.df.columns = ["TxnDate", "Narr", "Ref", "Dt_post", "Dr", "Cr", "Close",]
        
        #Remove debit entries
        self.df = self.df[np.isnan(self.df.Dr)]
        
        #Drop columns not needed
        self.df = self.df.drop(columns=['Dt_post', 'Dr', 'Close'])

        #Remove credit entries of Internal Transfer
        self.df = self.df[self.df.Narr.str.contains('INTERNAL TRANSFER') == False]

        #Remove NEFT returns
        self.df = self.df[self.df.Narr.str.contains('NEFT RETURN') == False]

        #Remove credit interest capitalized
        self.df = self.df[self.df.Narr.str.contains('CREDIT INTEREST CAPITALISED') == False]

        # Get payment mode
        self.df['Mode'] = self.df.Narr.apply(self.get_paymode)

        # Convert string dates to actual dates
        self.df.TxnDate = pd.to_datetime(self.df.TxnDate, format="%d/%m/%y")

        # Create yymm
        self.df['YYMM'] = str(pd.DatetimeIndex(self.df['TxnDate']).year) + "-" + str(pd.DatetimeIndex(self.df['TxnDate']).month)
        self.df['YYMM'] = pd.to_datetime(self.df['TxnDate']).dt.to_period('M')
        
        # Capture account type as column
        self.df['Acc Type'] = self.ac_type
        
        #Guess student name
        self.df['Student'] = self.df.Narr.apply(self.guess_student)

        #Sort by student name
        self.df = self.df.sort_values('Student')

        #Replace unidentified students name as UNIDENTIFIED
        self.df.Student =  self.df.Student.fillna(value="Unidentified")



    def get_paymode(self, text):
        if text.startswith('IMPS'):
            return 'IMPS'
        elif text.startswith('NEFT'):
            return 'NEFT'
        elif '-TPT-' in text:
            return 'TPT'
        elif text.startswith('UPI'):
            return 'UPI'
        elif 'NET BANKING SI' in text:
            return 'SI'
        elif 'CHQ DEP' in text or 'CHEQUE DEPOSIT' in text:
            return 'CHQ'
        elif 'CASH DEP' in text:
            return 'CASH'
        else:
            return ''

        
    def guess_student(self, text):
        pattern_m = {}
        pattern_m['XX - Sumit/Vidula Naikode'] = 'SHAILESH DAGADU'
        pattern_m['XX - Kali/Shri Shukla1'] = 'KALI SHRI'
        pattern_m['XX - Kali/Shri Shukla2'] = '15781130003652'
        pattern_m['XX - Kali/Shri Shukla3'] = 'VISHAL KIRIT'
        pattern_m['XX - Shreya/Rishi'] = 'SANDEEP MADHUKAR'
        pattern_m['XX - Riya/Siya'] = 'BHUPENDRA'
        pattern_m['XX - Vinayak/Shivraj1'] = 'GAJANAN SHANKAR'
        pattern_m['XX - Kshitija/Mayuresh Dhole1'] = 'CASH DEP JUINAGAR BRA'
        pattern_m['XX - Kshitija/Mayuresh Dhole2'] = 'KSHITIJABEAUTYCARE1@OKSBI'
        pattern_m['XX - Krishit/Nishi Pokar'] = 'RAJESH NARAN POKAR'

        pattern_m['XX - Donation1'] = 'DONATION'
        pattern_m['XX - Donation2'] = 'SAMREDHI SAREEN'
        pattern_m['XX - Donation3'] = 'KETAN SAHADEO'
        pattern_m['XX - Donation4'] = 'R R SARDA'
        pattern_m['PP - Aarohi Mishra'] = 'AAROHI ISHI'
        pattern_m['03 - Dev Madkaikar'] = 'MAST DEV'
        pattern_m['03 - Tashvi Vartak'] = 'TASHVI'
                
        pattern_m['04 - Joy Tanna'] = 'JOY'
        pattern_m['04 - Agrima Mishra1'] = 'AGRIMA'
        pattern_m['04 - Agrima Mishra2'] = 'MAHADEV'
        
        pattern_m['05 - Auro Arpan1'] = 'RAKESH KUMAR SAHOO'
        pattern_m['05 - Auro Arpan2'] = 'AURO ARPAN'
        pattern_m['05 - Auro Arpan3'] = '01331020001095'
        pattern_m['05 - Gaurav Chopra'] = 'RAHUL VINOD CHOPRA'
        pattern_m['05 - Rohit Nayak'] = 'RASHMI V NAYAK'
        
        pattern_m['06 - Isha Bapat'] = 'ISHA BAPAT'
        pattern_m['06 - Jeet Paswan'] = 'PASWAN'
        
        pattern_m['07 - Yash Gajra1'] = 'YASH GAJRA'
        pattern_m['07 - Yash Gajra2'] = 'AVG ASSOCIATES'
        pattern_m['07 - Harsh Ambesange'] = 'RAKHEE SUDHANSHU AMBESANGE'
        pattern_m['07 - Piyush Kale'] = 'SUREKHA  DNYANESHWAR KALE'
        pattern_m['07 - Jaivik Vyas'] = 'RENUKADARSHANVYAS'
        pattern_m['07 - Kaustubh Abnave'] = 'KAUSTUBH'
        pattern_m['07 - Saee Gujare'] = 'PRASHANT DATTU GUJAR'
        pattern_m['07 - Pari Mestry1'] = '15741000010660'
        pattern_m['07 - Pari Mestry2'] = 'AMOL N MESTRY'
        pattern_m['07 - Sarthak Sawant1'] = 'SARTHAK'
        pattern_m['07 - Sarthak Sawant2'] = '01851050034331'
        pattern_m['07 - Shivaji Bhegade1'] = 'SHREE DRONAGIREE AUSHADHALAY'
        pattern_m['07 - Shivaji Bhegade2'] = 'JSBP0000005-005220100015923'
        pattern_m['07 - Shivaji Bhegade3'] = 'BHEGADE'
        pattern_m['07 - Shlok Shewate'] = 'SHIVRAJ PACKAGING'

        pattern_m['08 - Aditya Mahajan'] = 'ANIL N MAHAJAN'
        pattern_m['08 - Arnav Sawant1'] = 'AARYAAMITSAWANT'
        pattern_m['08 - Arnav Sawant2'] = 'RAVINDRA R SAWANT'
        pattern_m['08 - Diya Pamnani'] = 'PALLAVI HIRANAND'
        pattern_m['08 - Rishi Agarwal1'] = 'MINU AGARWAL'
        pattern_m['08 - Rishi Agarwal2'] = 'RITU AGARWAL'
        pattern_m['08 - Yash Kshirsagar'] = 'VITTHAL BALASAHEB'
        pattern_m['08 - Yuvaraj Mudaliyar1'] = 'VIJAY RAMSWAMY MUDALIYAR'
        pattern_m['08 - Yuvaraj Mudaliyar2'] = 'VIJAY RAMSWAMY MUDAL'
        pattern_m['09 - Aditya Bhise'] = 'NITIN CHHAGAN BHISE'
        
        pattern_m['09 - Moulya1'] = 'MOULYA'
        pattern_m['09 - Moulya2'] = 'RAVINDRA A V'
        pattern_m['09 - Krushna Borhade'] = 'BORHADE'
        pattern_m['09 - Parth Kedia'] = 'GLOBAL ENGINEERING'
        
        pattern_m['10 - Arya Gawas'] = 'PRADIP GOVIND GAWAS'
        pattern_m['10 - Bhavya'] = 'BHAVYA PATEL'
        pattern_m['10 - Parth Agarwal'] = 'GANESH MAHAVIR PRASAD AGARWAL'        
        pattern_m['10 - Rucha Shelar'] = 'SANJAY S SHELAR'
        pattern_m['10 - Shubh Shah1'] = 'NAYAN VASANT SHAH'
        pattern_m['10 - Shubh Shah2'] = 'SHUBH TRADING'
        pattern_m['10 - Rasika Desai1'] = 'CHAITRA DESAI'
        pattern_m['10 - Rasika Desai2'] = '02271140016593'
        pattern_m['10 - Riya Patil1'] = 'SHAILESH S PATIL'
        pattern_m['10 - Riya Patil2'] = '00121250002434'
        pattern_m['10 - Sai Bagade'] = 'SANTOSH SHANTARAM'
        pattern_m['10 - Sakshi Tekawade'] = 'OCEAN'
        pattern_m['10 - Samruddhi Pawar1'] = 'AMOL JANAR'
        pattern_m['10 - Samruddhi Pawar2'] = 'AMOL PAWAR'
        pattern_m['10 - Siddhant Kamble'] = 'CHOPSTIX'
        pattern_m['10 - Soham Gawali'] = 'SURYAKANT JAGANNATH'
        pattern_m['10 - Ved Ahire1'] = 'VED AHIRE'
        pattern_m['10 - Ved Ahire2'] = '28021000006912'
        pattern_m['10 - Ved Ahire3'] = 'ASHWINIAHIRE'

        pattern_m['Past - Kukreja'] = 'KUKREJA'
        pattern_m['Past - Hariharan'] = 'HARIHARAN'
        pattern_m['Past - Sanavi'] = 'VISHWRAJ SHIVRAJ'
        pattern_m['Past - Utsav'] = 'ARVIND KUMAR GUPTA'
        pattern_m['Past - Soham Pokharkar'] = 'POKHARKAR'

        pattern_m['Maid Service Sapariya'] = '01491140000718'



# Laundry/maid/electricity expense
# DRVARSHAM@OKHDFCBANK
# ROHITTANNA@OKHDFCBANK
# RAJAN88@UPI



        for key, val in pattern_m.items():
            if val in text:
                return key
        return None
    
    def __str__(self):

        ret = '\n'.join([self.fname, self.ac_no, self.ac_type, self.date_from + ' - ' + self.date_to \
            , str(self.last_row), str(self.count_rows) \
            ] +  self.proc_status)
        return ret + '\n\n'


def dict_summary(df):
    return {"Filename": df.fname, \
        "AccNo": df.ac_no, \
        "AccType": df.ac_type, \
        "DateFrom": df.date_from,
        "DateTo":df.date_to,
        "Rows":df.count_rows}


def summarize(df1, df2, df3):
    # df is pandas data frame
    summ1 = pd.DataFrame(dict_summary(df1), index=[1])
    summ2 = pd.DataFrame(dict_summary(df2), index=[2])
    summ3 = pd.DataFrame(dict_summary(df3), index=[3])

    return pd.concat([summ1, summ2, summ3])
 



if __name__ == '__main__':
    
    #Set output filename
    output_fname = "output_20mar2020.xlsx"

    #Set input files
    df1 = fee_file("15mar2020_1.xls")
    df2 = fee_file("15mar2020_4.xls")
    df3 = fee_file("15mar2020_3.xls")

    # print(df1, df2, df3)
    df_summary = summarize(df1, df2, df3)

    df_all = pd.concat([df1.df, df2.df, df3.df])

    # If student name has a digit as suffix, remove that suffix
    df_all.Student = df_all.Student.replace(to_replace=r'(.*)\d$', value=r'\1', regex=True)

    # Pivot the data
    p1 = pd.pivot_table(df_all, index=["Acc Type", "Student"], \
        columns=["YYMM"], values=["Cr"], aggfunc=[np.sum], fill_value=0)
    
    #Output to excel - https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_excel.html
    with pd.ExcelWriter(output_fname) as writer:  # doctest: +SKIP
        df_summary.to_excel(writer, sheet_name='Input summary')
        df_all.to_excel(writer, sheet_name='Data', index=False)
        p1.to_excel(writer, sheet_name="Pivot")

    import openpyxl
    wb = openpyxl.load_workbook(output_fname)
    ws = wb['Pivot']
    ws.column_dimensions["A"].width=20
    ws.column_dimensions["B"].width=25
    from openpyxl.styles import Alignment, Font
    for i in range(5, 200, 1):
        thiscell = ws.cell(row=i, column=2)
        thiscell.alignment = Alignment(horizontal='left')
        if thiscell.value != 'Unidentified':
            thiscell.font = Font(bold=False)
    
    wb.active = ws
    wb.save(filename=output_fname)
    